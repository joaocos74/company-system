import customtkinter as ctk
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
import os
import datetime

# ====================
# POPUP DE MULTI-SELEÇÃO
# ====================
class MultiSelectPopup(ctk.CTkToplevel):
    def __init__(self, parent, titulo, itens, selecionados_iniciais=None):
        super().__init__(parent)
        self.title(titulo)
        self.geometry("400x300")
        self.selecionados_iniciais = selecionados_iniciais if selecionados_iniciais else []
        self.selected = []

        container = ctk.CTkFrame(self)
        container.pack(fill="both", expand=True, padx=10, pady=10)

        self.scrollable_frame = ctk.CTkScrollableFrame(container, width=350, height=200)
        self.scrollable_frame.pack(fill="both", expand=True)

        self.checkbox_vars = {}
        for item in itens:
            var = ctk.BooleanVar(value=item in self.selecionados_iniciais)
            chk = ctk.CTkCheckBox(self.scrollable_frame, text=item, variable=var)
            chk.pack(anchor="w", padx=5, pady=5)
            self.checkbox_vars[item] = var

        btn_confirmar = ctk.CTkButton(container, text="Confirmar", command=self.confirmar)
        btn_confirmar.pack(pady=10)

    def confirmar(self):
        self.selected = [
            item
            for item, var in self.checkbox_vars.items()
            if var.get() is True
        ]
        self.destroy()

# ====================
# JANELA PRINCIPAL
# ====================
class PesquisarApp(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Pesquisar Clientes")
        self.geometry("1000x700")
        self.resizable(True, True)

        # Armazena as classes selecionadas (do popup)
        self.selected_classes = []

        self.create_pesquisar()

    def create_pesquisar(self):
        container = ctk.CTkFrame(self)
        container.pack(fill=BOTH, expand=True, padx=20, pady=10)

        ctk.CTkLabel(
            container,
            text="PESQUISAR",
            font=("Century Gothic bold", 24),
            text_color="blue"
        ).pack(pady=10)

        # ==================
        # FRAME DE FILTROS
        # ==================
        filter_frame = ctk.CTkFrame(container)
        filter_frame.pack(pady=10, fill=X)

        self.id_value = StringVar()
        self.nivel_value = StringVar()
        self.rs_pf_value = StringVar()
        self.nome_fantasia_value = StringVar()
        self.endereco_value = StringVar()
        self.cnpj_cpf_value = StringVar()
        self.cnae_value = StringVar()
        self.parecer_value = StringVar()
        self.ultima_inspecao_value = StringVar()
        self.alvara_value = StringVar()
        self.vigi_risco_value = StringVar()
        self.observacao_value = StringVar()
        self.baixados_value = StringVar()
        self.excluidos_value = StringVar()

        # Carrega IDs e Classes do Excel
        self.carregar_opcoes()

        # Linha 0
        ctk.CTkLabel(filter_frame, text="ID:", font=("Century Gothic", 14)).grid(row=0, column=0, padx=5, pady=5, sticky=W)
        ctk.CTkComboBox(filter_frame, variable=self.id_value, values=self.id_options).grid(row=0, column=1, padx=5, pady=5, sticky=W)

        ctk.CTkLabel(filter_frame, text="NÍVEL:", font=("Century Gothic", 14)).grid(row=0, column=2, padx=5, pady=5, sticky=W)
        ctk.CTkComboBox(filter_frame, variable=self.nivel_value, values=["NIVEL 1", "NIVEL 2", "NIVEL 3"]).grid(row=0, column=3, padx=5, pady=5, sticky=W)

        # Linha 1
        ctk.CTkLabel(filter_frame, text="CLASSE(S):", font=("Century Gothic", 14)).grid(row=1, column=0, padx=5, pady=5, sticky=W)
        
        # Botão para abrir popup MultiSelectPopup
        self.btn_classe = ctk.CTkButton(
            filter_frame,
            text="Selecionar Classes",
            command=self.selecionar_classes
        )
        self.btn_classe.grid(row=1, column=1, padx=5, pady=5, sticky=W)

        # Label para mostrar as classes selecionadas
        self.lbl_classe_selecionadas = ctk.CTkLabel(filter_frame, text="", font=("Century Gothic", 12))
        self.lbl_classe_selecionadas.grid(row=1, column=2, padx=5, pady=5, sticky=W, columnspan=2)

        # Linha 2
        ctk.CTkLabel(filter_frame, text="RAZÃO SOCIAL/PF:", font=("Century Gothic", 14)).grid(row=2, column=0, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.rs_pf_value).grid(row=2, column=1, padx=5, pady=5, sticky=W)

        ctk.CTkLabel(filter_frame, text="NOME FANTASIA:", font=("Century Gothic", 14)).grid(row=2, column=2, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.nome_fantasia_value).grid(row=2, column=3, padx=5, pady=5, sticky=W)

        # Linha 3
        ctk.CTkLabel(filter_frame, text="ENDEREÇO:", font=("Century Gothic", 14)).grid(row=3, column=0, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.endereco_value).grid(row=3, column=1, padx=5, pady=5, sticky=W)

        ctk.CTkLabel(filter_frame, text="CNPJ/CPF:", font=("Century Gothic", 14)).grid(row=3, column=2, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.cnpj_cpf_value).grid(row=3, column=3, padx=5, pady=5, sticky=W)

        # Linha 4
        ctk.CTkLabel(filter_frame, text="CNAE (Principal):", font=("Century Gothic", 14)).grid(row=4, column=0, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.cnae_value).grid(row=4, column=1, padx=5, pady=5, sticky=W)

        ctk.CTkLabel(filter_frame, text="PARECER TÉCNICO:", font=("Century Gothic", 14)).grid(row=4, column=2, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.parecer_value).grid(row=4, column=3, padx=5, pady=5, sticky=W)

        # Linha 5
        ctk.CTkLabel(filter_frame, text="ÚLTIMA INSPEÇÃO:", font=("Century Gothic", 14)).grid(row=5, column=0, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.ultima_inspecao_value).grid(row=5, column=1, padx=5, pady=5, sticky=W)

        ctk.CTkLabel(filter_frame, text="ALVARÁ:", font=("Century Gothic", 14)).grid(row=5, column=2, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.alvara_value).grid(row=5, column=3, padx=5, pady=5, sticky=W)

        # Linha 6
        ctk.CTkLabel(filter_frame, text="VIGI-RISCO:", font=("Century Gothic", 14)).grid(row=6, column=0, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.vigi_risco_value).grid(row=6, column=1, padx=5, pady=5, sticky=W)

        ctk.CTkLabel(filter_frame, text="OBSERVAÇÕES:", font=("Century Gothic", 14)).grid(row=6, column=2, padx=5, pady=5, sticky=W)
        ctk.CTkEntry(filter_frame, textvariable=self.observacao_value).grid(row=6, column=3, padx=5, pady=5, sticky=W)

        # Linha 7
        ctk.CTkLabel(filter_frame, text="BAIXADOS:", font=("Century Gothic", 14)).grid(row=7, column=0, padx=5, pady=5, sticky=W)
        ctk.CTkComboBox(filter_frame, variable=self.baixados_value, values=["BAIXADO", "Não", ""]).grid(row=7, column=1, padx=5, pady=5, sticky=W)

        ctk.CTkLabel(filter_frame, text="EXCLUIDOS:", font=("Century Gothic", 14)).grid(row=7, column=2, padx=5, pady=5, sticky=W)
        ctk.CTkComboBox(filter_frame, variable=self.excluidos_value, values=["EXCLUÍDO", "Não", ""]).grid(row=7, column=3, padx=5, pady=5, sticky=W)

        # ==================
        # FRAME DE BOTÕES
        # ==================
        button_frame = ctk.CTkFrame(container)
        button_frame.pack(pady=10)

        ctk.CTkButton(button_frame, text="Pesquisar", command=self.pesquisar, fg_color="green").grid(row=0, column=0, padx=10, pady=5)
        ctk.CTkButton(button_frame, text="Limpar", command=self.limpar, fg_color="gray").grid(row=0, column=1, padx=10, pady=5)

        # >>> BOTÃO "SALVAR EXCEL" <<<
        ctk.CTkButton(button_frame, text="Salvar Excel", command=self.salvar_excel, fg_color="blue").grid(row=0, column=2, padx=10, pady=5)

        self.stats_label = ctk.CTkLabel(container, text="Resultados encontrados: 0", font=("Century Gothic", 14))
        self.stats_label.pack(pady=5)

        # =========================
        # FRAME DE RESULTADOS
        # =========================
        self.result_frame = ctk.CTkFrame(container)
        self.result_frame.pack(fill=BOTH, expand=True, pady=10)

        # >>> SCROLLBARS VERTICAL E HORIZONTAL <<<
        self.tree = ttk.Treeview(self.result_frame, xscrollcommand=None, yscrollcommand=None)

        # Scrollbar VERTICAL
        self.scrollbar_y = Scrollbar(self.result_frame, orient=VERTICAL, command=self.tree.yview)
        self.scrollbar_y.pack(side=RIGHT, fill=Y)

        # Scrollbar HORIZONTAL
        self.scrollbar_x = Scrollbar(self.result_frame, orient=HORIZONTAL, command=self.tree.xview)
        self.scrollbar_x.pack(side=BOTTOM, fill=X)

        # Conecta o Treeview às barras de rolagem
        self.tree.configure(
            yscrollcommand=self.scrollbar_y.set,
            xscrollcommand=self.scrollbar_x.set
        )
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)
        # >>> FIM SCROLLBARS <<<

    def carregar_opcoes(self):
        """Carrega as opções de ID e Classe a partir do Excel."""
        try:
            workbook = openpyxl.load_workbook("clientes.xlsx")
            folha = workbook.active

            self.id_options = []
            self.classe_options = []

            for row in folha.iter_rows(min_row=2, values_only=True):
                if row[0] not in (None, ""):
                    self.id_options.append(str(row[0]))
                if len(row) >= 3 and row[2] not in (None, ""):
                    self.classe_options.append(str(row[2]))

            self.id_options = sorted(set(self.id_options))
            self.classe_options = sorted(set(self.classe_options))

        except FileNotFoundError:
            self.id_options = []
            self.classe_options = []

    def selecionar_classes(self):
        """
        Abre o popup de seleção múltipla e salva o resultado em self.selected_classes
        """
        popup = MultiSelectPopup(
            self,
            titulo="Selecione as Classes",
            itens=self.classe_options,
            selecionados_iniciais=self.selected_classes
        )
        # Deixa modal até fechar
        self.wait_window(popup)

        # Atualiza a lista de classes selecionadas
        self.selected_classes = popup.selected

        # Exibe no label
        self.lbl_classe_selecionadas.configure(text=", ".join(self.selected_classes))

    def pesquisar(self):
        """
        Lê o arquivo Excel, filtra e exibe no Treeview.
        """
        try:
            workbook = openpyxl.load_workbook("clientes.xlsx")
            folha = workbook.active

            # Limpa o Treeview
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Define colunas a partir do cabeçalho do Excel
            cabecalho = [cell.value for cell in folha[1] if cell.value is not None]
            self.tree["columns"] = cabecalho
            self.tree["show"] = "headings"

            # Ajusta largura de colunas (exemplo)
            for col in cabecalho:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=150, anchor=W)

            total = 0
            for row in folha.iter_rows(min_row=2, values_only=True):
                if len(row) < len(cabecalho):
                    continue
                if self.filtrar(row):
                    linha_formatada = [str(valor) if valor is not None else "" for valor in row[:len(cabecalho)]]
                    self.tree.insert("", END, values=linha_formatada)
                    total += 1

            self.stats_label.configure(text=f"Resultados encontrados: {total}")

        except FileNotFoundError:
            messagebox.showerror("Erro", "O arquivo 'clientes.xlsx' não foi encontrado.")

    def filtrar(self, row):
        """
        Aplica os critérios de filtro em cada linha do Excel.
        Índices:
         0: ID
         1: Nível
         2: Classe
         3: Razão Social / PF
         ...
        14: Excluídos
        """
        if len(row) < 15:
            return False

        id_excel          = str(row[0]) if row[0] else ""
        nivel_excel       = str(row[1]) if row[1] else ""
        classe_excel      = str(row[2]) if row[2] else ""
        rs_pf_excel       = str(row[3]) if row[3] else ""
        nome_fant_excel   = str(row[4]) if row[4] else ""
        endereco_excel    = str(row[5]) if row[5] else ""
        cnpj_cpf_excel    = str(row[6]) if row[6] else ""
        cnae_excel        = str(row[7]) if row[7] else ""
        parecer_excel     = str(row[8]) if row[8] else ""
        ultima_insp_excel = row[9]
        alvara_excel      = str(row[10]) if row[10] else ""
        vigi_risco_excel  = str(row[11]) if row[11] else ""
        obs_excel         = str(row[12]) if row[12] else ""
        baixado_excel     = str(row[13]) if row[13] else ""
        excluido_excel    = str(row[14]) if row[14] else ""

        # Filtro por ID
        if self.id_value.get().strip():
            if id_excel != self.id_value.get().strip():
                return False

        # Filtro por Nível
        if self.nivel_value.get().strip():
            if nivel_excel != self.nivel_value.get().strip():
                return False

        # Filtro por Classes selecionadas
        if self.selected_classes:  # Se há pelo menos 1 classe na lista
            if classe_excel not in self.selected_classes:
                return False

        # Filtro por Razão Social/PF
        filtro_rs_pf = self.rs_pf_value.get().strip().lower()
        if filtro_rs_pf:
            if filtro_rs_pf not in rs_pf_excel.lower():
                return False

        # Filtro por Nome Fantasia
        filtro_nome = self.nome_fantasia_value.get().strip().lower()
        if filtro_nome:
            if filtro_nome not in nome_fant_excel.lower():
                return False

        # Filtro por Endereço
        filtro_end = self.endereco_value.get().strip().lower()
        if filtro_end:
            if filtro_end not in endereco_excel.lower():
                return False

        # Filtro por CNPJ/CPF
        filtro_cnpj_cpf = self.cnpj_cpf_value.get().strip().lower()
        if filtro_cnpj_cpf:
            if filtro_cnpj_cpf not in cnpj_cpf_excel.lower():
                return False

        # Filtro por CNAE
        filtro_cnae = self.cnae_value.get().strip().lower()
        if filtro_cnae:
            if filtro_cnae not in cnae_excel.lower():
                return False

        # Filtro por Parecer
        filtro_parecer = self.parecer_value.get().strip().lower()
        if filtro_parecer:
            if filtro_parecer not in parecer_excel.lower():
                return False

        # Filtro por Última inspeção
        filtro_inspecao = self.ultima_inspecao_value.get().strip().lower()
        if filtro_inspecao:
            # Se for data datetime.date ou datetime.datetime, converte para string
            if isinstance(ultima_insp_excel, (datetime.date, datetime.datetime)):
                inspecao_str = ultima_insp_excel.strftime("%d/%m/%Y")
            else:
                inspecao_str = str(ultima_insp_excel) if ultima_insp_excel else ""

            # Se o usuário digitou 4 dígitos (exemplo "2023"), verifica se consta na data
            if len(filtro_inspecao) == 4 and filtro_inspecao.isdigit():
                if filtro_inspecao not in inspecao_str:
                    return False
            else:
                if filtro_inspecao not in inspecao_str.lower():
                    return False

        # Filtro por Alvará
        filtro_alvara = self.alvara_value.get().strip().lower()
        if filtro_alvara:
            if filtro_alvara not in alvara_excel.lower():
                return False

        # Filtro por Vigi-Risco
        filtro_vigi = self.vigi_risco_value.get().strip().lower()
        if filtro_vigi:
            if filtro_vigi not in vigi_risco_excel.lower():
                return False

        # Filtro por Observações
        filtro_obs = self.observacao_value.get().strip().lower()
        if filtro_obs:
            if filtro_obs not in obs_excel.lower():
                return False

        # Filtro por Baixados
        if self.baixados_value.get().strip():
            if baixado_excel.lower() != self.baixados_value.get().strip().lower():
                return False

        # Filtro por Excluídos
        if self.excluidos_value.get().strip():
            if excluido_excel.lower() != self.excluidos_value.get().strip().lower():
                return False

        return True

    def limpar(self):
        """Limpa os campos e o Treeview."""
        self.id_value.set("")
        self.nivel_value.set("")
        self.rs_pf_value.set("")
        self.nome_fantasia_value.set("")
        self.endereco_value.set("")
        self.cnpj_cpf_value.set("")
        self.cnae_value.set("")
        self.parecer_value.set("")
        self.ultima_inspecao_value.set("")
        self.alvara_value.set("")
        self.vigi_risco_value.set("")
        self.observacao_value.set("")
        self.baixados_value.set("")
        self.excluidos_value.set("")

        # Limpa seleção de classes
        self.selected_classes = []
        self.lbl_classe_selecionadas.configure(text="")

        # Limpa o Treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.stats_label.configure(text="Resultados encontrados: 0")

    # >>> NOVA FUNÇÃO PARA SALVAR EM EXCEL <<<
    def salvar_excel(self):
        """
        Cria um novo arquivo Excel com o conteúdo do Treeview.
        """
        try:
            # Dialog para escolher local e nome do arquivo
            arquivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("Todos os Arquivos", "*.*")]
            )
            if not arquivo:  # Se o usuário cancelar
                return

            # Obtém as colunas e os dados do Treeview
            colunas = self.tree["columns"]

            # Cria uma nova pasta de trabalho do Excel
            new_wb = openpyxl.Workbook()
            ws = new_wb.active
            ws.title = "Pesquisa"

            # Escreve o cabeçalho
            for col_index, col_name in enumerate(colunas, start=1):
                ws.cell(row=1, column=col_index, value=col_name)

            # Escreve os dados das linhas
            row_excel = 2
            for item in self.tree.get_children():
                valores = self.tree.item(item, "values")
                for col_index, valor in enumerate(valores, start=1):
                    ws.cell(row=row_excel, column=col_index, value=valor)
                row_excel += 1

            # Salva o arquivo
            new_wb.save(arquivo)
            messagebox.showinfo("Sucesso", "Excel salvo com sucesso!")

        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao salvar o Excel:\n{e}")
