import customtkinter as ctk
from tkinter import messagebox, ttk
import openpyxl
import datetime
import unicodedata
from dateutil import parser as date_parser


def norm(s):
    if not s:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


class UsuarioApp(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Gerenciamento de Usuário")
        self.geometry("1200x850")
        self.minsize(1000, 700)

        self.arquivo = "cadastros.xlsx"
        self.responsavel = None
        self.dados = []
        self.responsaveis_lista = []

        self.carregar_dados()
        self.criar_ui()

    def carregar_dados(self):
        try:
            wb = openpyxl.load_workbook(self.arquivo, data_only=True)
            ws = wb.active

            # Encontrar colunas
            cabecalhos = {}
            for col, cell in enumerate(ws[1], 1):
                if not cell.value:
                    continue
                txt = norm(cell.value)
                if "id" in txt:
                    cabecalhos["id"] = col - 1
                elif "fantasia" in txt or ("nome" in txt and "responsavel" not in txt):
                    cabecalhos["nome"] = col - 1
                elif "responsavel" in txt:
                    cabecalhos["responsavel"] = col - 1
                elif "inspe" in txt:
                    cabecalhos["inspecao"] = col - 1

            self.col_id = cabecalhos.get("id", 0)
            self.col_nome = cabecalhos.get("nome", 1)
            self.col_responsavel = cabecalhos.get("responsavel", 2)
            self.col_inspecao = cabecalhos.get("inspecao", 3)

            # Carregar dados
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                estab_id = row[self.col_id].value
                nome = row[self.col_nome].value
                responsavel = row[self.col_responsavel].value
                inspecao = row[self.col_inspecao].value

                if estab_id and nome:
                    self.dados.append({
                        "id": estab_id,
                        "nome": str(nome).strip(),
                        "responsavel": str(responsavel).strip() if responsavel else None,
                        "inspecao": inspecao,
                        "row": row_idx,
                    })

            wb.close()

            # Lista de responsáveis únicos
            self.responsaveis_lista = sorted(list(set([
                d["responsavel"] for d in self.dados if d["responsavel"]
            ])))

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar: {e}")
            self.destroy()

    def criar_ui(self):
        # Main scroll frame
        main_scroll = ctk.CTkScrollableFrame(self, fg_color="#f8f8f8")
        main_scroll.pack(fill="both", expand=True, padx=0, pady=0)

        # TITULO
        titulo = ctk.CTkLabel(
            main_scroll,
            text="Gerenciamento de Usuário",
            font=("Segoe UI", 28, "bold"),
            text_color="#054721"
        )
        titulo.pack(pady=20, padx=20)

        # ====== FRAME SELEÇÃO ======
        frame_sel = ctk.CTkFrame(main_scroll, fg_color="white", corner_radius=12)
        frame_sel.pack(fill="x", padx=20, pady=10)

        lbl_sel = ctk.CTkLabel(
            frame_sel,
            text="Selecione ou Crie um Responsável",
            font=("Segoe UI", 16, "bold"),
            text_color="#054721"
        )
        lbl_sel.pack(pady=(15, 10), anchor="w", padx=20)

        # Combo
        f1 = ctk.CTkFrame(frame_sel, fg_color="transparent")
        f1.pack(pady=8, padx=20)

        self.combo = ctk.CTkComboBox(
            f1,
            values=self.responsaveis_lista or ["Nenhum"],
            state="readonly" if self.responsaveis_lista else "disabled",
            width=300,
            height=40,
            font=("Segoe UI", 13)
        )
        self.combo.pack(side="left", padx=5)
        if self.responsaveis_lista:
            self.combo.set("Selecione...")

        btn_sel = ctk.CTkButton(
            f1,
            text="Selecionar",
            width=130,
            height=40,
            command=self.on_selecionar,
            fg_color="#054721",
            hover_color="#043618",
            font=("Segoe UI", 13, "bold")
        )
        btn_sel.pack(side="left", padx=5)

        # OU
        sep = ctk.CTkLabel(frame_sel, text="ou", text_color="gray", font=("Segoe UI", 12))
        sep.pack(pady=5)

        # Novo
        f2 = ctk.CTkFrame(frame_sel, fg_color="transparent")
        f2.pack(pady=8, padx=20)

        self.entry_novo = ctk.CTkEntry(
            f2,
            placeholder_text="Digite o nome do novo responsável",
            width=300,
            height=40,
            font=("Segoe UI", 13)
        )
        self.entry_novo.pack(side="left", padx=5)

        btn_novo = ctk.CTkButton(
            f2,
            text="Criar Novo",
            width=130,
            height=40,
            command=self.on_criar_novo,
            fg_color="#2c7a4f",
            hover_color="#1f5739",
            font=("Segoe UI", 13, "bold")
        )
        btn_novo.pack(side="left", padx=5)

        ctk.CTkLabel(frame_sel, text="").pack(pady=5)

        self.main_scroll = main_scroll

    def on_selecionar(self):
        resp = self.combo.get()
        if not resp or resp == "Selecione...":
            messagebox.showwarning("Aviso", "Selecione um responsável")
            return
        self.responsavel = resp
        self.mostrar_dados()

    def on_criar_novo(self):
        nome = self.entry_novo.get().strip()
        if not nome:
            messagebox.showwarning("Aviso", "Digite o nome")
            return
        if nome in self.responsaveis_lista:
            messagebox.showwarning("Aviso", "Responsável já existe")
            return

        self.responsaveis_lista.append(nome)
        self.responsaveis_lista.sort()
        self.combo.configure(values=self.responsaveis_lista)
        self.entry_novo.delete(0, "end")
        self.responsavel = nome
        self.mostrar_dados()

    def mostrar_dados(self):
        # Remove frame anterior se existir
        for widget in self.main_scroll.winfo_children():
            if widget != self.main_scroll.winfo_children()[0]:  # Pula o título
                widget.destroy()

        # Re-criar título
        titulo = ctk.CTkLabel(
            self.main_scroll,
            text="Gerenciamento de Usuário",
            font=("Segoe UI", 28, "bold"),
            text_color="#054721"
        )
        titulo.pack(pady=20, padx=20)

        # Re-criar frame seleção (com dados)
        self.criar_frame_selecao()

        # Calcular dados
        vinculados = [d for d in self.dados if d["responsavel"] == self.responsavel]
        total = len(vinculados)

        ano_atual = datetime.datetime.now().year
        inspecionados = 0

        for e in vinculados:
            try:
                if e["inspecao"]:
                    data = e["inspecao"]
                    if isinstance(data, datetime.datetime):
                        if data.year == ano_atual:
                            inspecionados += 1
                    elif isinstance(data, datetime.date):
                        if data.year == ano_atual:
                            inspecionados += 1
                    else:
                        ano_str = str(data)[-4:]
                        if ano_str.isdigit() and int(ano_str) == ano_atual:
                            inspecionados += 1
            except:
                pass

        cobertura = (inspecionados / total * 100) if total > 0 else 0

        # ====== FRAME ESTATÍSTICAS ======
        frame_stats = ctk.CTkFrame(self.main_scroll, fg_color="white", corner_radius=12)
        frame_stats.pack(fill="x", padx=20, pady=10)

        # Título com botão excluir
        f_titulo = ctk.CTkFrame(frame_stats, fg_color="transparent")
        f_titulo.pack(fill="x", padx=20, pady=(15, 10))

        lbl = ctk.CTkLabel(
            f_titulo,
            text=f"Estatísticas - {self.responsavel}",
            font=("Segoe UI", 18, "bold"),
            text_color="#054721"
        )
        lbl.pack(side="left", expand=True)

        btn_del = ctk.CTkButton(
            f_titulo,
            text="🗑 Excluir Responsável",
            width=180,
            height=35,
            command=self.on_excluir,
            fg_color="#c62828",
            hover_color="#8e0000",
            font=("Segoe UI", 12, "bold")
        )
        btn_del.pack(side="right")

        # Cards
        f_cards = ctk.CTkFrame(frame_stats, fg_color="transparent")
        f_cards.pack(fill="x", padx=20, pady=15)

        # Card 1 - Total
        card1 = ctk.CTkFrame(f_cards, fg_color="#2c7a4f", corner_radius=12)
        card1.pack(side="left", fill="both", expand=True, padx=10, pady=5)

        c1_1 = ctk.CTkLabel(card1, text="Total de Estabelecimentos", text_color="white", font=("Segoe UI", 13))
        c1_1.pack(pady=(15, 5), anchor="w", padx=15)

        c1_2 = ctk.CTkLabel(card1, text=str(total), text_color="white", font=("Segoe UI", 40, "bold"))
        c1_2.pack(anchor="w", padx=15, pady=(5, 15))

        # Card 2 - Inspecionados
        card2 = ctk.CTkFrame(f_cards, fg_color="#1565C0", corner_radius=12)
        card2.pack(side="left", fill="both", expand=True, padx=10, pady=5)

        c2_1 = ctk.CTkLabel(card2, text=f"Inspecionados em {ano_atual}", text_color="white", font=("Segoe UI", 13))
        c2_1.pack(pady=(15, 5), anchor="w", padx=15)

        c2_2 = ctk.CTkLabel(card2, text=str(inspecionados), text_color="white", font=("Segoe UI", 40, "bold"))
        c2_2.pack(anchor="w", padx=15, pady=(5, 15))

        # Card 3 - Cobertura
        card3 = ctk.CTkFrame(f_cards, fg_color="#E65100", corner_radius=12)
        card3.pack(side="left", fill="both", expand=True, padx=10, pady=5)

        c3_1 = ctk.CTkLabel(card3, text="Cobertura", text_color="white", font=("Segoe UI", 13))
        c3_1.pack(pady=(15, 5), anchor="w", padx=15)

        c3_2 = ctk.CTkLabel(card3, text=f"{cobertura:.1f}%", text_color="white", font=("Segoe UI", 40, "bold"))
        c3_2.pack(anchor="w", padx=15, pady=(5, 15))

        # ====== FRAME PERÍODO ======
        frame_periodo = ctk.CTkFrame(self.main_scroll, fg_color="white", corner_radius=12)
        frame_periodo.pack(fill="x", padx=20, pady=10)

        lbl_per = ctk.CTkLabel(
            frame_periodo,
            text="Inspeções em Período Personalizado",
            font=("Segoe UI", 16, "bold"),
            text_color="#054721"
        )
        lbl_per.pack(pady=(15, 10), anchor="w", padx=20)

        f_per = ctk.CTkFrame(frame_periodo, fg_color="transparent")
        f_per.pack(padx=20, pady=(0, 15), fill="x")

        ctk.CTkLabel(f_per, text="Data Inicial:", text_color="#054721", font=("Segoe UI", 13)).pack(side="left", padx=(0, 8))

        self.entry_data1 = ctk.CTkEntry(f_per, placeholder_text="01/01/2025", width=120, height=35, font=("Segoe UI", 12))
        self.entry_data1.pack(side="left", padx=5)

        ctk.CTkLabel(f_per, text="Data Final:", text_color="#054721", font=("Segoe UI", 13)).pack(side="left", padx=(15, 8))

        self.entry_data2 = ctk.CTkEntry(f_per, placeholder_text="10/02/2025", width=120, height=35, font=("Segoe UI", 12))
        self.entry_data2.pack(side="left", padx=5)

        btn_buscar = ctk.CTkButton(
            f_per,
            text="Buscar",
            width=100,
            height=35,
            command=self.on_buscar_periodo,
            fg_color="#054721",
            hover_color="#043618",
            font=("Segoe UI", 12, "bold")
        )
        btn_buscar.pack(side="left", padx=10)

        self.lbl_resultado = ctk.CTkLabel(f_per, text="", text_color="#E65100", font=("Segoe UI", 13, "bold"))
        self.lbl_resultado.pack(side="left", padx=15)

        # ====== FRAME GERENCIAMENTO ======
        frame_ger = ctk.CTkFrame(self.main_scroll, fg_color="white", corner_radius=12)
        frame_ger.pack(fill="both", expand=True, padx=20, pady=10)

        lbl_ger = ctk.CTkLabel(
            frame_ger,
            text="Gerenciar Estabelecimentos",
            font=("Segoe UI", 16, "bold"),
            text_color="#054721"
        )
        lbl_ger.pack(pady=(15, 10), anchor="w", padx=20)

        # Duas colunas
        f_cols = ctk.CTkFrame(frame_ger, fg_color="transparent")
        f_cols.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.criar_coluna_esquerda(f_cols, vinculados)
        self.criar_coluna_direita(f_cols, vinculados)

    def criar_frame_selecao(self):
        frame_sel = ctk.CTkFrame(self.main_scroll, fg_color="white", corner_radius=12)
        frame_sel.pack(fill="x", padx=20, pady=10)

        lbl_sel = ctk.CTkLabel(
            frame_sel,
            text="Selecione ou Crie um Responsável",
            font=("Segoe UI", 16, "bold"),
            text_color="#054721"
        )
        lbl_sel.pack(pady=(15, 10), anchor="w", padx=20)

        f1 = ctk.CTkFrame(frame_sel, fg_color="transparent")
        f1.pack(pady=8, padx=20)

        self.combo = ctk.CTkComboBox(
            f1,
            values=self.responsaveis_lista or ["Nenhum"],
            state="readonly" if self.responsaveis_lista else "disabled",
            width=300,
            height=40,
            font=("Segoe UI", 13)
        )
        self.combo.pack(side="left", padx=5)
        if self.responsaveis_lista:
            self.combo.set(self.responsavel)

        btn_sel = ctk.CTkButton(
            f1,
            text="Selecionar",
            width=130,
            height=40,
            command=self.on_selecionar,
            fg_color="#054721",
            hover_color="#043618",
            font=("Segoe UI", 13, "bold")
        )
        btn_sel.pack(side="left", padx=5)

        sep = ctk.CTkLabel(frame_sel, text="ou", text_color="gray", font=("Segoe UI", 12))
        sep.pack(pady=5)

        f2 = ctk.CTkFrame(frame_sel, fg_color="transparent")
        f2.pack(pady=8, padx=20)

        self.entry_novo = ctk.CTkEntry(
            f2,
            placeholder_text="Digite o nome do novo responsável",
            width=300,
            height=40,
            font=("Segoe UI", 13)
        )
        self.entry_novo.pack(side="left", padx=5)

        btn_novo = ctk.CTkButton(
            f2,
            text="Criar Novo",
            width=130,
            height=40,
            command=self.on_criar_novo,
            fg_color="#2c7a4f",
            hover_color="#1f5739",
            font=("Segoe UI", 13, "bold")
        )
        btn_novo.pack(side="left", padx=5)

        ctk.CTkLabel(frame_sel, text="").pack(pady=5)

    def criar_coluna_esquerda(self, parent, vinculados):
        f = ctk.CTkFrame(parent, fg_color="#f5f5f5", corner_radius=10)
        f.pack(side="left", fill="both", expand=True, padx=(0, 10))

        titulo = ctk.CTkLabel(f, text="Estabelecimentos Disponíveis", font=("Segoe UI", 14, "bold"), text_color="#054721")
        titulo.pack(pady=(12, 8), anchor="w", padx=12)

        f_tree = ctk.CTkFrame(f, fg_color="white", corner_radius=8)
        f_tree.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        scrollbar = ttk.Scrollbar(f_tree)
        scrollbar.pack(side="right", fill="y")

        self.tree1 = ttk.Treeview(f_tree, columns=("ID", "Nome", "Insp"), show="headings", height=12, yscrollcommand=scrollbar.set)
        self.tree1.heading("ID", text="ID")
        self.tree1.heading("Nome", text="Estabelecimento")
        self.tree1.heading("Insp", text=f"Insp. {datetime.datetime.now().year}")
        self.tree1.column("ID", width=45, anchor="center")
        self.tree1.column("Nome", width=200)
        self.tree1.column("Insp", width=70, anchor="center")
        scrollbar.config(command=self.tree1.yview)
        self.tree1.pack(fill="both", expand=True)

        # Dados disponíveis
        ids_vinc = set([d["id"] for d in vinculados])
        idx = 0
        for d in self.dados:
            if d["id"] not in ids_vinc:
                status = "✓" if self.verificar_inspecao(d) else "✗"
                tag = "odd" if idx % 2 == 0 else "even"
                self.tree1.insert("", "end", values=(d["id"], d["nome"], status), tags=(tag, str(d["id"])))
                idx += 1

        self.tree1.tag_configure("odd", background="#fafafa")
        self.tree1.tag_configure("even", background="white")

        btn = ctk.CTkButton(f, text="➜ Vincular Selecionados", command=self.on_vincular, fg_color="#2c7a4f", hover_color="#1f5739", font=("Segoe UI", 13, "bold"), height=40)
        btn.pack(padx=12, pady=8, fill="x")

    def criar_coluna_direita(self, parent, vinculados):
        f = ctk.CTkFrame(parent, fg_color="#f5f5f5", corner_radius=10)
        f.pack(side="left", fill="both", expand=True, padx=(10, 0))

        titulo = ctk.CTkLabel(f, text=f"Vinculados a {self.responsavel}", font=("Segoe UI", 14, "bold"), text_color="#054721")
        titulo.pack(pady=(12, 8), anchor="w", padx=12)

        f_tree = ctk.CTkFrame(f, fg_color="white", corner_radius=8)
        f_tree.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        scrollbar = ttk.Scrollbar(f_tree)
        scrollbar.pack(side="right", fill="y")

        self.tree2 = ttk.Treeview(f_tree, columns=("ID", "Nome", "Insp"), show="headings", height=12, yscrollcommand=scrollbar.set)
        self.tree2.heading("ID", text="ID")
        self.tree2.heading("Nome", text="Estabelecimento")
        self.tree2.heading("Insp", text=f"Insp. {datetime.datetime.now().year}")
        self.tree2.column("ID", width=45, anchor="center")
        self.tree2.column("Nome", width=200)
        self.tree2.column("Insp", width=70, anchor="center")
        scrollbar.config(command=self.tree2.yview)
        self.tree2.pack(fill="both", expand=True)

        # Dados vinculados
        idx = 0
        for d in vinculados:
            status = "✓" if self.verificar_inspecao(d) else "✗"
            tag = "odd" if idx % 2 == 0 else "even"
            self.tree2.insert("", "end", values=(d["id"], d["nome"], status), tags=(tag, str(d["id"])))
            idx += 1

        self.tree2.tag_configure("odd", background="#fafafa")
        self.tree2.tag_configure("even", background="white")

        btn = ctk.CTkButton(f, text="✗ Desvincular Selecionados", command=self.on_desvincular, fg_color="#c62828", hover_color="#8e0000", font=("Segoe UI", 13, "bold"), height=40)
        btn.pack(padx=12, pady=8, fill="x")

    def verificar_inspecao(self, estab):
        ano_atual = datetime.datetime.now().year
        try:
            if estab["inspecao"]:
                data = estab["inspecao"]
                if isinstance(data, datetime.datetime):
                    return data.year == ano_atual
                elif isinstance(data, datetime.date):
                    return data.year == ano_atual
                else:
                    ano_str = str(data)[-4:]
                    return ano_str.isdigit() and int(ano_str) == ano_atual
        except:
            pass
        return False

    def on_buscar_periodo(self):
        try:
            d1 = self.entry_data1.get().strip()
            d2 = self.entry_data2.get().strip()

            if not d1 or not d2:
                messagebox.showwarning("Aviso", "Preencha as datas")
                return

            data_ini = date_parser.parse(d1, dayfirst=True)
            data_fim = date_parser.parse(d2, dayfirst=True)

            if data_ini > data_fim:
                messagebox.showwarning("Aviso", "Data inicial > final")
                return

            vinculados = [d for d in self.dados if d["responsavel"] == self.responsavel]
            contagem = 0

            for e in vinculados:
                try:
                    if e["inspecao"]:
                        data = e["inspecao"]
                        if isinstance(data, (datetime.datetime, datetime.date)):
                            d_insp = data if isinstance(data, datetime.date) else data.date()
                            if data_ini.date() <= d_insp <= data_fim.date():
                                contagem += 1
                except:
                    pass

            total = len(vinculados)
            perc = (contagem / total * 100) if total > 0 else 0
            resultado = f"Período {d1} a {d2} → {contagem} estabelecimentos ({perc:.1f}%)"
            self.lbl_resultado.configure(text=resultado)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro: {e}")

    def on_vincular(self):
        selecionados = self.tree1.selection()
        if not selecionados:
            messagebox.showwarning("Aviso", "Selecione itens")
            return

        try:
            wb = openpyxl.load_workbook(self.arquivo)
            ws = wb.active

            for item in selecionados:
                estab_id = int(self.tree1.item(item)["values"][0])
                for d in self.dados:
                    if d["id"] == estab_id:
                        ws.cell(row=d["row"], column=self.col_responsavel + 1, value=self.responsavel)
                        d["responsavel"] = self.responsavel

            wb.save(self.arquivo)
            wb.close()
            self.mostrar_dados()
            messagebox.showinfo("Sucesso", f"{len(selecionados)} vinculado(s)")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro: {e}")

    def on_desvincular(self):
        selecionados = self.tree2.selection()
        if not selecionados:
            messagebox.showwarning("Aviso", "Selecione itens")
            return

        if not messagebox.askyesno("Confirmar", f"Desvincular {len(selecionados)} item(ns)?"):
            return

        try:
            wb = openpyxl.load_workbook(self.arquivo)
            ws = wb.active

            for item in selecionados:
                estab_id = int(self.tree2.item(item)["values"][0])
                for d in self.dados:
                    if d["id"] == estab_id:
                        ws.cell(row=d["row"], column=self.col_responsavel + 1, value=None)
                        d["responsavel"] = None

            wb.save(self.arquivo)
            wb.close()
            self.mostrar_dados()
            messagebox.showinfo("Sucesso", f"{len(selecionados)} desvinculado(s)")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro: {e}")

    def on_excluir(self):
        if not messagebox.askyesno("Confirmar", f"Excluir '{self.responsavel}'?\nIsso desvinculará todos os estabelecimentos."):
            return

        try:
            wb = openpyxl.load_workbook(self.arquivo)
            ws = wb.active

            for d in self.dados:
                if d["responsavel"] == self.responsavel:
                    ws.cell(row=d["row"], column=self.col_responsavel + 1, value=None)
                    d["responsavel"] = None

            wb.save(self.arquivo)
            wb.close()

            self.responsaveis_lista.remove(self.responsavel)
            self.responsavel = None
            self.combo.configure(values=self.responsaveis_lista or ["Nenhum"])

            # Limpar interface
            for widget in self.main_scroll.winfo_children()[1:]:
                widget.destroy()

            messagebox.showinfo("Sucesso", "Responsável excluído")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro: {e}")
