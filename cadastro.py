import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib

class CadastroApp(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Cadastro de Clientes")
        self.geometry("1000x600")
        self.resizable(True, True)

        # Tema inicial
        self.appearance_mode = "System"
        ctk.set_appearance_mode(self.appearance_mode)

        # Menu para mudar o tema
        self.menu = Menu(self)
        self.config(menu=self.menu)
        theme_menu = Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="Tema", menu=theme_menu)
        theme_menu.add_command(label="Claro", command=lambda: self.change_theme("Light"))
        theme_menu.add_command(label="Escuro", command=lambda: self.change_theme("Dark"))
        theme_menu.add_command(label="Sistema", command=lambda: self.change_theme("System"))

        self.create_cadastro()

    def create_cadastro(self):
        container = ctk.CTkFrame(self)
        container.pack(fill=BOTH, expand=True, padx=20, pady=20)

        # Título
        ctk.CTkLabel(container, text="CADASTRO", font=("Century Gothic bold", 24), text_color="teal").grid(row=0, column=0, columnspan=4, pady=10)

        # Variáveis
        self.id_value = StringVar()
        self.nivel_value = StringVar()
        self.classe_value = StringVar()
        self.razao_social_value = StringVar()
        self.nome_fantasia_value = StringVar()
        self.endereco_value = StringVar()
        self.cnpj_cpf_value = StringVar()
        self.cnae_value = StringVar()
        self.parecer_value = StringVar()
        self.inspecao_value = StringVar()
        self.alvara_value = StringVar()
        self.vigi_risco_value = StringVar()
        self.observacoes_value = StringVar()
        self.baixados_value = StringVar()
        self.excluidos_value = StringVar()

        # Cria arquivo Excel se não existir
        ficheiro = pathlib.Path("clientes.xlsx")
        if not ficheiro.exists():
            workbook = openpyxl.Workbook()
            folha = workbook.active
            headers = ["ID", "NIVEL", "CLASSE", "RAZÃO SOCIAL OU PESSOA FÍSICA", "NOME FANTASIA", "ENDEREÇO", "CNPJ OU CPF",
                       "CNAE (Principal)", "Nº PARECER TÉCNICO", "ÚLTIMA INSPEÇÃO", "ALVARÁ", "VIGI-RISCO",
                       "OBSERVAÇÕES", "BAIXADOS", "EXCLUIDOS"]
            folha.append(headers)
            workbook.save("clientes.xlsx")

        # Campos e Labels
        fields = [
            ("ID", self.id_value, None),
            ("Nível", self.nivel_value, ["NIVEL 1", "NIVEL 2", "NIVEL 3"]),
            ("Classe", self.classe_value, None),
            ("Razão Social/Pessoa Física", self.razao_social_value, None),
            ("Nome Fantasia", self.nome_fantasia_value, None),
            ("Endereço", self.endereco_value, None),
            ("CNPJ ou CPF", self.cnpj_cpf_value, None),
            ("CNAE (Principal)", self.cnae_value, None),
            ("Nº Parecer Técnico", self.parecer_value, None),
            ("Última Inspeção (texto livre)", self.inspecao_value, None),
            ("Alvará (DD/MM/AAAA)", self.alvara_value, None),
            ("Vigi-Risco (DD/MM/AAAA)", self.vigi_risco_value, None),
            ("Observações", self.observacoes_value, None),
            ("Baixados", self.baixados_value, ["BAIXADO", "Não"]),
            ("Excluídos", self.excluidos_value, ["EXCLUÍDO", "Não"]),
        ]

        # Configuração em duas colunas
        for i, (label, variable, values) in enumerate(fields):
            row = i // 2 + 1
            col = (i % 2) * 2
            ctk.CTkLabel(container, text=label + ":", font=("Century Gothic bold", 14)).grid(row=row, column=col, padx=10, pady=5, sticky=W)
            if label == "ID":
                ctk.CTkEntry(container, textvariable=variable, width=300, state="disabled").grid(row=row, column=col + 1, padx=10, pady=5, sticky=W)
            elif values:
                ctk.CTkComboBox(container, values=values, variable=variable).grid(row=row, column=col + 1, padx=10, pady=5, sticky=W)
            elif label == "Observações":
                self.observacoes_textbox = ctk.CTkTextbox(container, width=300, height=50)
                self.observacoes_textbox.grid(row=row, column=col + 1, padx=10, pady=5, sticky=W)
            else:
                ctk.CTkEntry(container, textvariable=variable, width=300).grid(row=row, column=col + 1, padx=10, pady=5, sticky=W)

        # Botões
        ctk.CTkButton(container, text="Salvar", command=self.salvar, fg_color="green").grid(row=len(fields)//2 + 2, column=0, columnspan=2, padx=10, pady=20)
        ctk.CTkButton(container, text="Limpar", command=self.limpar, fg_color="gray").grid(row=len(fields)//2 + 2, column=2, columnspan=2, padx=10, pady=20)

    def salvar(self):
        confirmar = messagebox.askyesno("Confirmação", "Deseja realmente salvar os dados?")
        if not confirmar:
            return

        workbook = openpyxl.load_workbook("clientes.xlsx")
        folha = workbook.active

        # Geração automática do ID
        id_auto = folha.max_row
        self.id_value.set(str(id_auto))

        # Obtém dados
        nivel = self.nivel_value.get()
        classe = self.classe_value.get()
        razao_social = self.razao_social_value.get()
        nome_fantasia = self.nome_fantasia_value.get()
        endereco = self.endereco_value.get()
        cnpj_cpf = self.cnpj_cpf_value.get()
        cnae = self.cnae_value.get()
        parecer = self.parecer_value.get()
        inspecao = self.inspecao_value.get()
        alvara = self.alvara_value.get()
        vigi_risco = self.vigi_risco_value.get()
        observacoes = self.observacoes_textbox.get("1.0", "end").strip()
        baixados = self.baixados_value.get()
        excluidos = self.excluidos_value.get()

        # Salva no Excel
        folha.append([id_auto, nivel, classe, razao_social, nome_fantasia, endereco, cnpj_cpf, cnae, parecer, inspecao, alvara, vigi_risco, observacoes, baixados, excluidos])
        workbook.save("clientes.xlsx")
        messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")

        self.limpar()

    def limpar(self):
        self.id_value.set("")
        self.nivel_value.set("")
        self.classe_value.set("")
        self.razao_social_value.set("")
        self.nome_fantasia_value.set("")
        self.endereco_value.set("")
        self.cnpj_cpf_value.set("")
        self.cnae_value.set("")
        self.parecer_value.set("")
        self.inspecao_value.set("")
        self.alvara_value.set("")
        self.vigi_risco_value.set("")
        self.observacoes_textbox.delete("1.0", "end")
        self.baixados_value.set("")
        self.excluidos_value.set("")

    def change_theme(self, mode):
        self.appearance_mode = mode
        ctk.set_appearance_mode(self.appearance_mode)
