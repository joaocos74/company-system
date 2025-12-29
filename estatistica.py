
import customtkinter as ctk
from tkinter import *
from tkinter import ttk, messagebox
import openpyxl
import re
import datetime
import unicodedata

# ========================
#  Estatística de Inspeção
# ========================
#
# Requisitos atendidos:
# - Layout mais bonito e organizado (cards de KPI + duas tabelas lado a lado)
# - Combo de ANO mostra o intervalo completo de anos [min..máximo encontrado]
# - Mostra "Total de estabelecimentos cadastrados" (geral) e cobertura (%)
# - Percentuais por NIVEL e por CLASSE com totais e inspecionados no ano
#
# Colunas esperadas (case/acento-insensível):
#   - 'NIVEL'   (ou 'NÍVEL')
#   - 'CLASSE'
#   - 'ÚLTIMA INSPEÇÃO' (qualquer variação contendo 'inspe' será aceita como fallback)

def _norm(s: str) -> str:
    """Normaliza string para comparação: lowercase + sem acento + sem espaços extras."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s

class EstatisticaApp(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Estatísticas de Inspeção")
        self.geometry("1120x720")
        self.minsize(980, 640)

        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")

        self.ano_value = StringVar()

        self._build_ui()
        self._safe_bootstrap()

    # ---------------- UI ----------------
    def _build_ui(self):
        self.container = ctk.CTkFrame(self, corner_radius=12)
        self.container.pack(fill=BOTH, expand=True, padx=16, pady=16)

        # Header
        header = ctk.CTkFrame(self.container, corner_radius=12, fg_color=("#05531f", "#05531f"))
        header.pack(fill=X, padx=0, pady=(0,12))
        ctk.CTkLabel(header, text="Painel de Estatísticas", font=("Segoe UI Semibold", 24),
                     text_color="white").pack(side=LEFT, padx=16, pady=12)

        # Filtros
        filtros = ctk.CTkFrame(self.container, corner_radius=12)
        filtros.pack(fill=X, pady=(0,12))

        ctk.CTkLabel(filtros, text="Ano da Inspeção", font=("Segoe UI", 14)).grid(row=0, column=0, padx=(16,8), pady=12, sticky=W)
        self.ano_combo = ctk.CTkComboBox(filtros, variable=self.ano_value, values=[],
                                         width=160, command=lambda _=None: self._calcular())
        self.ano_combo.grid(row=0, column=1, padx=(0,12), pady=12, sticky=W)

        self.btn_recarregar = ctk.CTkButton(filtros, text="Recarregar", width=120, command=self._recarregar_anos)
        self.btn_recarregar.grid(row=0, column=2, padx=(0,12), pady=12)

        self.lbl_hint = ctk.CTkLabel(filtros, text="", font=("Segoe UI", 12), text_color="gray")
        self.lbl_hint.grid(row=0, column=3, padx=8, pady=12, sticky=W)

        filtros.grid_columnconfigure(10, weight=1)  # espaço flexível

        # KPIs
        kpis = ctk.CTkFrame(self.container, corner_radius=12)
        kpis.pack(fill=X, pady=(0,12))

        self.card_total = self._kpi_card(kpis, "Total de Estabelecimentos (Geral)", "—")
        self.card_ano   = self._kpi_card(kpis, "Inspecionados no Ano Selecionado", "—")
        self.card_cob   = self._kpi_card(kpis, "Cobertura no Ano (%)", "—")

        self.card_total.pack(side=LEFT, fill=X, expand=True, padx=(0,8), pady=8)
        self.card_ano.pack(side=LEFT, fill=X, expand=True, padx=8, pady=8)
        self.card_cob.pack(side=LEFT, fill=X, expand=True, padx=(8,0), pady=8)

        # Tabelas
        tables = ctk.CTkFrame(self.container, corner_radius=12)
        tables.pack(fill=BOTH, expand=True, pady=(0,0))

        # Por Nível
        left = ctk.CTkFrame(tables, corner_radius=12)
        left.pack(side=LEFT, fill=BOTH, expand=True, padx=(0,8), pady=0)

        ctk.CTkLabel(left, text="Percentual por Nível", font=("Segoe UI Semibold", 16)).pack(anchor=W, padx=12, pady=(12,4))
        self.tree_nivel = self._make_tree(left, cols=(
            ("nivel","Nível",160),
            ("total","Total do Nível",140),
            ("inspecionados","Inspecionados no Ano",180),
            ("percentual","Percentual (%)",140),
        ))
        self.tree_nivel.pack(fill=BOTH, expand=True, padx=12, pady=(0,12))

        # Por Classe
        right = ctk.CTkFrame(tables, corner_radius=12)
        right.pack(side=LEFT, fill=BOTH, expand=True, padx=(8,0), pady=0)

        ctk.CTkLabel(right, text="Percentual por Classe", font=("Segoe UI Semibold", 16)).pack(anchor=W, padx=12, pady=(12,4))
        self.tree_classe = self._make_tree(right, cols=(
            ("classe","Classe",220),
            ("total","Total da Classe",140),
            ("inspecionados","Inspecionados no Ano",180),
            ("percentual","Percentual (%)",140),
        ))
        self.tree_classe.pack(fill=BOTH, expand=True, padx=12, pady=(0,12))

    def _kpi_card(self, parent, title, value):
        card = ctk.CTkFrame(parent, corner_radius=16)
        title_lbl = ctk.CTkLabel(card, text=title, font=("Segoe UI", 13), text_color="gray70")
        title_lbl.pack(anchor=W, padx=16, pady=(12,0))
        value_lbl = ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 26))
        value_lbl.pack(anchor=W, padx=16, pady=(0,12))
        card.value_lbl = value_lbl
        return card

    def _make_tree(self, parent, cols):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except:
            pass
        style.configure("Treeview", rowheight=28, font=("Segoe UI", 11))
        style.configure("Treeview.Heading", font=("Segoe UI Semibold", 11))

        cols_ids = [c[0] for c in cols]
        tree = ttk.Treeview(parent, columns=cols_ids, show="headings", selectmode="browse")
        for cid, title, width in cols:
            tree.heading(cid, text=title)
            tree.column(cid, width=width, anchor=W, stretch=True)

        yscroll = ttk.Scrollbar(parent, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=yscroll.set)

        # pack later by caller; but yscroll must be on same parent
        tree.bind("<Configure>", lambda e: yscroll.place_forget())
        tree.bind("<Expose>", lambda e: yscroll.place(relx=1.0, rely=0, relheight=1.0, anchor="ne"))
        return tree

    # ---------------- Bootstrapping ----------------
    def _safe_bootstrap(self):
        try:
            self._recarregar_anos()
        except FileNotFoundError:
            messagebox.showerror("Erro", "O arquivo 'cadastros.xlsx' não foi encontrado.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao iniciar:\n{e}")

    # ---------------- Dados ----------------
    def _recarregar_anos(self):
        anos = self._carregar_anos()  # set/list
        if not anos:
            self.ano_combo.configure(values=[])
            self.ano_value.set("")
            self.lbl_hint.configure(text="Nenhum ano encontrado na planilha.")
            return

        anos_ordenados = sorted(list(anos))
        # Intervalo completo [min..max]
        ano_min, ano_max = anos_ordenados[0], anos_ordenados[-1]
        intervalo = [str(a) for a in range(ano_min, ano_max + 1)]
        self.ano_combo.configure(values=intervalo)
        self.ano_value.set(str(ano_max))  # sempre seleciona o último ano disponível
        self.lbl_hint.configure(text=f"Intervalo disponível: {ano_min}–{ano_max}")
        self._calcular()

    def _carregar_anos(self):
        wb = openpyxl.load_workbook("cadastros.xlsx", data_only=True)
        sh = wb.active
        col_insp = self._col_index(sh, "ÚLTIMA INSPEÇÃO")
        anos = set()
        for r in sh.iter_rows(min_row=2, values_only=True):
            val = r[col_insp] if col_insp < len(r) else None
            ano = self._extrair_ano(val)
            if ano is not None:
                anos.add(ano)
        return anos

    def _col_index(self, sheet, header_name):
        """Retorna o índice (0-based) da coluna cujo cabeçalho bate com header_name (acento/case-insensível).
           Fallback: primeira coluna que contenha 'inspe' para a coluna de inspeção."""
        target = _norm(header_name)
        headers = [c.value for c in sheet[1]]
        # tentativa direta
        for idx, h in enumerate(headers):
            if _norm(h) == target:
                return idx
        # fallback para inspeção
        if "inspe" in target:
            for idx, h in enumerate(headers):
                if "inspe" in _norm(h):
                    return idx
        # fallback genérico para nivel/classe com startswith
        for idx, h in enumerate(headers):
            if _norm(h).startswith(target):
                return idx
        raise ValueError(f"Coluna '{header_name}' não encontrada. Cabeçalhos: {headers}")

    def _extrair_ano(self, valor):
        if isinstance(valor, (datetime.datetime, datetime.date)):
            return valor.year
        if valor is None:
            return None
        s = str(valor)
        m = re.search(r'(19|20)\d{2}', s)
        if m:
            return int(m.group(0))
        return None

    def _calcular(self):
        ano_str = self.ano_value.get().strip()
        if not ano_str:
            return
        try:
            ano = int(ano_str)
        except:
            messagebox.showerror("Erro", "Ano inválido. Selecione um ano na lista.")
            return

        try:
            wb = openpyxl.load_workbook("cadastros.xlsx", data_only=True)
            sh = wb.active
        except FileNotFoundError:
            messagebox.showerror("Erro", "O arquivo 'cadastros.xlsx' não foi encontrado.")
            return

        # Mapeia colunas principais (com fallback de acentos)
        try:
            idx_nivel  = self._col_index(sh, "NIVEL")
        except:
            idx_nivel  = self._col_index(sh, "NÍVEL")
        idx_classe = self._col_index(sh, "CLASSE")
        idx_insp   = self._col_index(sh, "ÚLTIMA INSPEÇÃO")

        total_cadastrados = 0
        total_inspecionados_ano = 0

        total_por_nivel = {}
        insp_por_nivel = {}
        total_por_classe = {}
        insp_por_classe = {}

        for r in sh.iter_rows(min_row=2, values_only=True):
            # conta cadastrado quando houver ao menos algum dado significativo
            if any(cell not in (None, "", " ") for cell in r):
                total_cadastrados += 1

            if len(r) <= max(idx_nivel, idx_classe, idx_insp):
                continue

            nivel = str(r[idx_nivel]).strip() if r[idx_nivel] not in (None, "") else ""
            classe = str(r[idx_classe]).strip() if r[idx_classe] not in (None, "") else ""
            ano_r = self._extrair_ano(r[idx_insp])

            if nivel:
                total_por_nivel[nivel] = total_por_nivel.get(nivel, 0) + 1
            if classe:
                total_por_classe[classe] = total_por_classe.get(classe, 0) + 1

            if ano_r == ano:
                total_inspecionados_ano += 1
                if nivel:
                    insp_por_nivel[nivel] = insp_por_nivel.get(nivel, 0) + 1
                if classe:
                    insp_por_classe[classe] = insp_por_classe.get(classe, 0) + 1

        # KPIs
        cobertura = (total_inspecionados_ano / total_cadastrados * 100) if total_cadastrados > 0 else 0.0
        self.card_total.value_lbl.configure(text=f"{total_cadastrados:,}".replace(",", "."))
        self.card_ano.value_lbl.configure(text=f"{total_inspecionados_ano:,}".replace(",", "."))
        self.card_cob.value_lbl.configure(text=f"{cobertura:.2f}%")

        # Atualiza tabelas
        for t in (self.tree_nivel, self.tree_classe):
            for item in t.get_children():
                t.delete(item)

        # Nível: ordena por nome
        for nivel in sorted(total_por_nivel.keys(), key=lambda x: (x is None, str(x))):
            tot = total_por_nivel.get(nivel, 0)
            insp = insp_por_nivel.get(nivel, 0)
            perc = (insp / tot * 100) if tot > 0 else 0.0
            self.tree_nivel.insert("", END, values=(nivel, tot, insp, f"{perc:.2f}"))

        # Classe: ordena por maior percentual, depois classe
        linhas_classe = []
        for classe, tot in total_por_classe.items():
            insp = insp_por_classe.get(classe, 0)
            perc = (insp / tot * 100) if tot > 0 else 0.0
            linhas_classe.append((classe, tot, insp, perc))
        linhas_classe.sort(key=lambda x: (-x[3], str(x[0])))
        for classe, tot, insp, perc in linhas_classe:
            self.tree_classe.insert("", END, values=(classe, tot, insp, f"{perc:.2f}"))

if __name__ == "__main__":
    # Para testes isolados
    root = ctk.CTk()
    root.withdraw()
    EstatisticaApp(root)
    root.mainloop()
