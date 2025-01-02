import customtkinter as ctk
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import openpyxl

class AtualizarApp(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Editar Clientes")
        self.geometry("1000x700")
        self.resizable(True, True)
        self.create_widgets()

    def create_widgets(self):
        container = ctk.CTkFrame(self)
        container.pack(fill=BOTH, expand=True, padx=20, pady=10)

        # Título
        ctk.CTkLabel(
            container,
            text="Editar Clientes",
            font=("Century Gothic bold", 24),
            text_color="blue"
        ).pack(pady=10)

        # Campos de entrada organizados em duas colunas
        form_frame = ctk.CTkFrame(container)
        form_frame.pack(pady=10, fill=X)

        self.fields = {
            "ID": StringVar(),
            "NÍVEL": StringVar(),
            "CLASSE": StringVar(),
            "RAZÃO SOCIAL/PF": StringVar(),
            "NOME FANTASIA": StringVar(),
            "ENDEREÇO": StringVar(),
            "CNPJ/CPF": StringVar(),
            "CNAE": StringVar(),
            "PARECER": StringVar(),
            "INSPEÇÃO": StringVar(),
            "ALVARÁ": StringVar(),
            "VIGI-RISCO": StringVar(),
            "OBSERVAÇÃO": StringVar(),
            "BAIXADOS": StringVar(),
            "EXCLUÍDOS": StringVar()
        }

        self.combobox_fields = ["NÍVEL", "CLASSE", "BAIXADOS", "EXCLUÍDOS"]

        self.combobox_widgets = {}

        for idx, (label, var) in enumerate(self.fields.items()):
            col = idx % 2
            row = idx // 2

            ctk.CTkLabel(form_frame, text=label + ":", font=("Century Gothic", 14)).grid(row=row, column=col * 2, padx=5, pady=5, sticky=W)

            if label in self.combobox_fields:
                combobox = ctk.CTkComboBox(form_frame, variable=var, values=[])
                combobox.grid(row=row, column=(col * 2) + 1, padx=5, pady=5, sticky=W)
                self.combobox_widgets[label] = combobox
            elif label == "ID":
                ctk.CTkEntry(form_frame, textvariable=var, state="disabled").grid(row=row, column=(col * 2) + 1, padx=5, pady=5, sticky=W)
            else:
                ctk.CTkEntry(form_frame, textvariable=var).grid(row=row, column=(col * 2) + 1, padx=5, pady=5, sticky=W)

        # Botões
        button_frame = ctk.CTkFrame(container)
        button_frame.pack(pady=10)

        ctk.CTkButton(button_frame, text="Limpar", command=self.limpar_campos, fg_color="gray").grid(row=0, column=0, padx=10, pady=5)
        ctk.CTkButton(button_frame, text="Pesquisar", command=self.pesquisar, fg_color="green").grid(row=0, column=1, padx=10, pady=5)
        ctk.CTkButton(button_frame, text="Alterar", command=self.alterar_registro, fg_color="orange").grid(row=0, column=2, padx=10, pady=5)

        # Tabela
        self.tree = ttk.Treeview(container, columns=list(self.fields.keys()), show="headings")
        self.tree.pack(fill=BOTH, expand=True, pady=10)

        for col in self.fields.keys():
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150)

        self.tree.bind("<Double-1>", self.carregar_dados)

        # Scrollbar
        scrollbar = Scrollbar(self.tree, orient=VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)

        # Carregar dados
        self.carregar_dados_excel()

    def carregar_dados_excel(self):
        try:
            self.workbook = openpyxl.load_workbook("clientes.xlsx")
            self.sheet = self.workbook.active

            # Preencher os valores dos ComboBox
            combobox_values = {key: set() for key in self.combobox_fields}

            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                for idx, (key, var) in enumerate(self.fields.items()):
                    if key in self.combobox_fields and row[idx] is not None:
                        combobox_values[key].add(str(row[idx]))

            for key, values in combobox_values.items():
                if key in self.combobox_widgets:
                    self.combobox_widgets[key].configure(values=list(sorted(values)))

            for row in self.tree.get_children():
                self.tree.delete(row)

            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", END, values=row)

        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo 'clientes.xlsx' não encontrado.")

    def carregar_dados(self, event):
        selected_item = self.tree.focus()
        if not selected_item:
            return

        values = self.tree.item(selected_item, "values")

        for idx, var in enumerate(self.fields.values()):
            var.set(values[idx] if idx < len(values) else "")

    def pesquisar(self):
        try:
            filtro = {label: var.get().strip() for label, var in self.fields.items() if var.get().strip()}

            self.workbook = openpyxl.load_workbook("clientes.xlsx")
            self.sheet = self.workbook.active

            for row in self.tree.get_children():
                self.tree.delete(row)

            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {label: row[idx] if idx < len(row) else "" for idx, label in enumerate(self.fields.keys())}

                if all(str(row_dict[key]).lower().find(str(value).lower()) != -1 for key, value in filtro.items()):
                    self.tree.insert("", END, values=row)

        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo 'clientes.xlsx' não encontrado.")

    def alterar_registro(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showwarning("Aviso", "Nenhum registro selecionado.")
            return

        confirm = messagebox.askyesno("Confirmar", "Deseja realmente alterar o registro?")
        if not confirm:
            return

        values = [var.get() for var in self.fields.values()]
        id_to_update = values[0]  # ID é o primeiro valor

        # Localiza a linha com o ID correspondente no Excel
        for row_idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            if str(row[0]) == id_to_update:
                for col_idx, value in enumerate(values, start=1):
                    if col_idx == 1:  # Não alterar o ID
                        continue
                    self.sheet.cell(row=row_idx, column=col_idx, value=value)
                break
        else:
            messagebox.showerror("Erro", "Registro não encontrado no Excel.")
            return

        self.workbook.save("clientes.xlsx")
        messagebox.showinfo("Sucesso", "Registro alterado com sucesso!")

        # Atualizar a visualização da tabela
        self.tree.item(selected_item, values=values)

    def limpar_campos(self):
        for var in self.fields.values():
            var.set("")
